# -*- coding: utf-8 -*-
#
# Copyright (c) 2019 Jürgen Mülbert. All rights reserved.
#
# Licensed under the EUPL, Version 1.2 or – as soon they
# will be approved by the European Commission - subsequent
# versions of the EUPL (the "Licence");
# You may not use this work except in compliance with the
# Licence.
# You may obtain a copy of the Licence at:
#
# https://joinup.ec.europa.eu/page/eupl-text-11-12
#
# Unless required by applicable law or agreed to in
# writing, software distributed under the Licence is
# distributed on an "AS IS" basis,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either
# express or implied.
# See the Licence for the specific language governing
# permissions and limitations under the Licence.
#
# Lizenziert unter der EUPL, Version 1.2 oder - sobald
#  diese von der Europäischen Kommission genehmigt wurden -
# Folgeversionen der EUPL ("Lizenz");
# Sie dürfen dieses Werk ausschließlich gemäß
# dieser Lizenz nutzen.
# Eine Kopie der Lizenz finden Sie hier:
#
# https://joinup.ec.europa.eu/page/eupl-text-11-12
#
# Sofern nicht durch anwendbare Rechtsvorschriften
# gefordert oder in schriftlicher Form vereinbart, wird
# die unter der Lizenz verbreitete Software "so wie sie
# ist", OHNE JEGLICHE GEWÄHRLEISTUNG ODER BEDINGUNGEN -
# ausdrücklich oder stillschweigend - verbreitet.
# Die sprachspezifischen Genehmigungen und Beschränkungen
# unter der Lizenz sind dem Lizenztext zu entnehmen.
#
"""
This Program will read a csv that contents service persons as csv.

Then reads an csv with informations of this people and split
the csv. for each person writes a seperate excel-file.

The Format of the person file is:
    | Name | Firstname |

The Format of the data file is:
    - Auftrag Nummer                pos 0 (String)
    - Hauptbereich                  pos 1 (String)
    - Auftragsdatum                 pos 2 (Datum)
    - Tage offen                    pos 3 (Int)
    - Deb.-Nr.                      pos 4 (String)
    - Deb.-Name                     pos 5 (String)
    - Verkäufer Serviceberater      pos 6 (String)
    - Arbeitswert                   pos 7 (float)
    - Teile                         pos 8 (float)
    - Fremdleistung                 pos 9 (float)
    - Andere                        pos 10 (float)
    - Gesamt                        pos 11 (float)
    - Auftragswert bereit geliefert pos 12 (float)
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Font

from ..core.logger import logger


class GenerateOrders:
    """Generate the orders for output."""

    def __init__(self, destdir):
        """Init the GenerateOrders Class."""
        self.dest_name = ""
        self.dest_dir = destdir

    def create(self, actual_name, actual_content):
        """Put all the data for the actual_name to the excel-file."""
        row_num = 1
        col_num = 1

        # Create a workbook and add a worksheet.
        if self.dest_dir:
            self.dest_name = os.path.join(
                os.path.abspath(
                    self.dest_dir,
                ), actual_name + ".xlsx",
            )
        else:
            self.dest_name = actual_name + ".xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = actual_name
        sheet.sheet_properties.tabColor = "1072BA"

        line_count = 0
        for print_line in actual_content:
            logger.debug(print_line)
            if "Auftrag Nr." in print_line[0]:
                # Write Header
                for item in print_line:
                    logger.debug("Header: " + item)
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = item
                    cell.font = Font(name="Courier", size=12)
                    cell.style = "Title"
                    cell.number_format = "text"
                    col_num += 1

                row_num += 1
                col_num = 1

            # Alle Zeilen mit dem BERATER Namen in die Exceldatei schreiben.
            if actual_name in print_line[6]:
                line_count += 1
                for item in print_line:
                    logger.debug(
                        "Name: {0} Data: {1} Count: {2}".formmat(
                            actual_name, item, line_count)
                    )
                    logger.debug(
                        "row_num: " + str(row_num) +
                        " col_num: " + str(col_num)
                    )

                    # Tage offen ist eine ganze Zahl
                    if col_num == 3:
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.value = item
                        cell.number_format = "dd.mm.yyyy"
                    # Tage offen ist eine ganze Zahl
                    elif col_num == 4:
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.value = float(item)
                        cell.number_format = "#,##0.00"
                    # Alles was nach Deb-Name ist, ist eine reale Zahl
                    elif col_num > 7:
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.value = float(item)
                        cell.number_format = "#,##0.00_€"
                    else:
                        sheet.cell(row=row_num, column=col_num).value = item

                    col_num += 1

                # Alle Daten für den aktuellen Auftrag geschrieben ->
                # nächste Zeile
                row_num += 1
                # und wieder ganz nach links.
                col_num = 1

        # Alles fertig Excel-Sheet schließen!
        workbook.save(self.dest_name)
        # Contains the name no lines then delete the file
        if line_count == 0:
            os.remove(self.dest_name)
